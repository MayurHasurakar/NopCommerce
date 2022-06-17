import openpyxl

def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_row)

def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_column)

def readData(file, sheetName, rows,col):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rows,column= col).value

def writeData(file, sheetName, rows, col,Data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row= rows, column=col).value = Data
    workbook.save(file)

















